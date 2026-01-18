from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_required, current_user
from extensions import db
from models.account_invoice import InvoiceHongKong, InvoiceThai
from models.customer import Customer
from datetime import datetime
from decimal import Decimal
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os

# Create blueprint
account_report = Blueprint('account_report', __name__, url_prefix='/account-report')

# ============================================
# INVOICES LIST & EXPORT
# ============================================

@account_report.route('/invoices')
@login_required
def invoices_list():
    """Display invoices with type filter"""
    # Get invoice type from query parameter (default: booking)
    invoice_type = request.args.get('type', 'booking')
    
    all_invoices = []
    
    if invoice_type == 'booking':
        # Get invoices from booking invoices table
        from models.invoice import Invoice
        booking_invoices = Invoice.query.order_by(Invoice.created_at.desc()).all()
        
        for inv in booking_invoices:
            all_invoices.append({
                'type': 'Booking',
                'id': inv.id,
                'customer_id': inv.booking_id,
                'cust_name': inv.title or '-',
                'company_name': '-',
                'total_pax': '-',
                'arrival_date': inv.invoice_date.strftime('%Y-%m-%d') if inv.invoice_date else '-',
                'departure_date': inv.due_date.strftime('%Y-%m-%d') if inv.due_date else '-',
                'guest_list': '',
                'flight_info': '',
                'air_ticket_cost': 0,
                'transportation_fee': 0,
                'advance_expense': 0,
                'tour_fee': 0,
                'vat': 0,
                'withholding_tax': 0,
                'total_tour_fee': float(inv.total_amount) if inv.total_amount else 0,
                'created_at': inv.created_at.isoformat() if inv.created_at else ''
            })
    
    elif invoice_type == 'hongkong':
        # Get Hong Kong invoices
        hk_invoices = InvoiceHongKong.query.order_by(InvoiceHongKong.created_at.desc()).all()
        for inv in hk_invoices:
            data = inv.to_dict()
            data['type'] = 'HK'
            all_invoices.append(data)
    
    elif invoice_type == 'thai':
        # Get Thai invoices
        thai_invoices = InvoiceThai.query.order_by(InvoiceThai.created_at.desc()).all()
        for inv in thai_invoices:
            data = inv.to_dict()
            data['type'] = 'Thai'
            all_invoices.append(data)
    
    return render_template('account_report/invoices_list.html', 
                         invoices=all_invoices, 
                         invoice_type=invoice_type)


@account_report.route('/invoices/export-excel')
@login_required
def export_invoices_excel():
    """Export invoices to Excel based on type filter"""
    # Get invoice type from query parameter (default: booking)
    invoice_type = request.args.get('type', 'booking')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Define headers
    headers = [
        'Type', 'ID', 'Customer ID', 'Customer Name', 'Total Pax', 
        'Arrival Date', 'Departure Date', 'Guest List', 'Flight Info',
        'Air Ticket Cost', 'Transportation Fee', 'Advance Expense', 
        'Tour Fee', 'VAT', 'Withholding Tax', 'Total Tour Fee'
    ]
    
    if invoice_type == 'booking':
        ws.title = "Booking Invoices"
        headers = [
            'Type', 'ID', 'Booking ID', 'Title', 'Invoice Number',
            'Invoice Date', 'Due Date', 'Total Amount', 'Status', 'Created'
        ]
    elif invoice_type == 'hongkong':
        ws.title = "Hong Kong Invoices"
    else:
        ws.title = "Thai Invoices"
    
    # Style for headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row = 2
    
    if invoice_type == 'booking':
        # Export Booking Invoices
        from models.invoice import Invoice
        booking_invoices = Invoice.query.order_by(Invoice.created_at.desc()).all()
        
        for inv in booking_invoices:
            ws.cell(row=row, column=1, value='Booking')
            ws.cell(row=row, column=2, value=inv.id)
            ws.cell(row=row, column=3, value=inv.booking_id)
            ws.cell(row=row, column=4, value=inv.title or '')
            ws.cell(row=row, column=5, value=inv.invoice_number or '')
            ws.cell(row=row, column=6, value=inv.invoice_date.strftime('%Y-%m-%d') if inv.invoice_date else '')
            ws.cell(row=row, column=7, value=inv.due_date.strftime('%Y-%m-%d') if inv.due_date else '')
            ws.cell(row=row, column=8, value=float(inv.total_amount) if inv.total_amount else 0.00)
            ws.cell(row=row, column=9, value=inv.status or '')
            ws.cell(row=row, column=10, value=inv.created_at.strftime('%Y-%m-%d') if inv.created_at else '')
            row += 1
    
    elif invoice_type == 'hongkong':
        # Export HK invoices
        hk_invoices = InvoiceHongKong.query.all()
        for inv in hk_invoices:
            ws.cell(row=row, column=1, value='Hong Kong')
            ws.cell(row=row, column=2, value=inv.id)
            ws.cell(row=row, column=3, value=inv.customer_id)
            ws.cell(row=row, column=4, value=inv.cust_name)
            ws.cell(row=row, column=5, value=inv.total_pax)
            ws.cell(row=row, column=6, value=inv.arrival_date.strftime('%Y-%m-%d') if inv.arrival_date else '')
            ws.cell(row=row, column=7, value=inv.departure_date.strftime('%Y-%m-%d') if inv.departure_date else '')
            ws.cell(row=row, column=8, value=inv.guest_list or '')
            ws.cell(row=row, column=9, value=inv.flight_info or '')
            ws.cell(row=row, column=10, value=float(inv.air_ticket_cost) if inv.air_ticket_cost else 0.00)
            ws.cell(row=row, column=11, value=float(inv.transportation_fee) if inv.transportation_fee else 0.00)
            ws.cell(row=row, column=12, value=float(inv.advance_expense) if inv.advance_expense else 0.00)
            ws.cell(row=row, column=13, value=float(inv.tour_fee) if inv.tour_fee else 0.00)
            ws.cell(row=row, column=14, value=float(inv.vat) if inv.vat else 0.00)
            ws.cell(row=row, column=15, value=float(inv.withholding_tax) if inv.withholding_tax else 0.00)
            ws.cell(row=row, column=16, value=float(inv.total_tour_fee) if inv.total_tour_fee else 0.00)
            row += 1
    
    else:  # thai
        # Export Thai invoices
        thai_invoices = InvoiceThai.query.all()
        for inv in thai_invoices:
            ws.cell(row=row, column=1, value='Thai')
            ws.cell(row=row, column=2, value=inv.id)
            ws.cell(row=row, column=3, value=inv.customer_id)
            ws.cell(row=row, column=4, value=inv.cust_name)
            ws.cell(row=row, column=5, value=inv.total_pax)
            ws.cell(row=row, column=6, value=inv.arrival_date.strftime('%Y-%m-%d') if inv.arrival_date else '')
            ws.cell(row=row, column=7, value=inv.departure_date.strftime('%Y-%m-%d') if inv.departure_date else '')
            ws.cell(row=row, column=8, value=inv.guest_list or '')
            ws.cell(row=row, column=9, value=inv.flight_info or '')
            ws.cell(row=row, column=10, value=float(inv.air_ticket_cost) if inv.air_ticket_cost else 0.00)
            ws.cell(row=row, column=11, value=float(inv.transportation_fee) if inv.transportation_fee else 0.00)
            ws.cell(row=row, column=12, value=float(inv.advance_expense) if inv.advance_expense else 0.00)
            ws.cell(row=row, column=13, value=float(inv.tour_fee) if inv.tour_fee else 0.00)
            ws.cell(row=row, column=14, value=float(inv.vat) if inv.vat else 0.00)
            ws.cell(row=row, column=15, value=float(inv.withholding_tax) if inv.withholding_tax else 0.00)
            ws.cell(row=row, column=16, value=float(inv.total_tour_fee) if inv.total_tour_fee else 0.00)
            row += 1
    
    # Apply borders to all cells
    max_col = 10 if invoice_type == 'booking' else 16
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=max_col):
        for cell in row_cells:
            cell.border = border
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp and type
    type_name = invoice_type.capitalize()
    filename = f"{type_name}_invoices_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ============================================
# HONG KONG INVOICE ROUTES
# ============================================

@account_report.route('/hongkong-invoice')
@login_required
def hongkong_invoice_list():
    """List all Hong Kong invoices"""
    invoices = InvoiceHongKong.query.order_by(InvoiceHongKong.created_at.desc()).all()
    return render_template('account_report/hongkong_invoice_list.html', invoices=invoices)


@account_report.route('/hongkong-invoice/create', methods=['GET', 'POST'])
@login_required
def hongkong_invoice_create():
    """Create new Hong Kong invoice"""
    if request.method == 'POST':
        try:
            # Parse dates
            arrival_date = None
            departure_date = None
            
            if request.form.get('arrival_date'):
                arrival_date = datetime.strptime(request.form.get('arrival_date'), '%Y-%m-%d').date()
            
            if request.form.get('departure_date'):
                departure_date = datetime.strptime(request.form.get('departure_date'), '%Y-%m-%d').date()
            
            # Create new invoice
            invoice = InvoiceHongKong(
                customer_id=int(request.form.get('customer_id')) if request.form.get('customer_id') else None,
                cust_name=request.form.get('cust_name'),
                company_name=request.form.get('company_name'),
                company_address=request.form.get('company_address'),
                company_tel=request.form.get('company_tel'),
                company_taxid=request.form.get('company_taxid'),
                company_contact=request.form.get('company_contact'),
                total_pax=int(request.form.get('total_pax', 0)),
                arrival_date=arrival_date,
                departure_date=departure_date,
                guest_list=request.form.get('guest_list'),
                flight_info=request.form.get('flight_info'),
                air_ticket_cost=Decimal(request.form.get('air_ticket_cost', 0)),
                transportation_fee=Decimal(request.form.get('transportation_fee', 0)),
                advance_expense=Decimal(request.form.get('advance_expense', 0)),
                tour_fee=Decimal(request.form.get('tour_fee', 0)),
                vat=Decimal(request.form.get('vat', 0)),
                withholding_tax=Decimal(request.form.get('withholding_tax', 0)),
                total_tour_fee=Decimal(request.form.get('total_tour_fee', 0))
            )
            
            db.session.add(invoice)
            db.session.commit()
            
            flash('Hong Kong Invoice created successfully!', 'success')
            return redirect(url_for('account_report.hongkong_invoice_view', invoice_id=invoice.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating invoice: {str(e)}', 'danger')
            return redirect(url_for('account_report.hongkong_invoice_create'))
    
    # GET request - show form
    customers = Customer.query.order_by(Customer.name).all()
    return render_template('account_report/hongkong_invoice_form.html', customers=customers, invoice=None)


@account_report.route('/hongkong-invoice/<int:invoice_id>')
@login_required
def hongkong_invoice_view(invoice_id):
    """View Hong Kong invoice"""
    invoice = InvoiceHongKong.query.get_or_404(invoice_id)
    return render_template('account_report/hongkong_invoice_view.html', invoice=invoice)


@account_report.route('/hongkong-invoice/<int:invoice_id>/edit', methods=['GET', 'POST'])
@login_required
def hongkong_invoice_edit(invoice_id):
    """Edit Hong Kong invoice"""
    invoice = InvoiceHongKong.query.get_or_404(invoice_id)
    
    if request.method == 'POST':
        try:
            # Parse dates
            if request.form.get('arrival_date'):
                invoice.arrival_date = datetime.strptime(request.form.get('arrival_date'), '%Y-%m-%d').date()
            
            if request.form.get('departure_date'):
                invoice.departure_date = datetime.strptime(request.form.get('departure_date'), '%Y-%m-%d').date()
            
            # Update fields
            invoice.customer_id = int(request.form.get('customer_id')) if request.form.get('customer_id') else None
            invoice.cust_name = request.form.get('cust_name')
            invoice.company_name = request.form.get('company_name')
            invoice.company_address = request.form.get('company_address')
            invoice.company_tel = request.form.get('company_tel')
            invoice.company_taxid = request.form.get('company_taxid')
            invoice.company_contact = request.form.get('company_contact')
            invoice.total_pax = int(request.form.get('total_pax', 0))
            invoice.guest_list = request.form.get('guest_list')
            invoice.flight_info = request.form.get('flight_info')
            invoice.air_ticket_cost = Decimal(request.form.get('air_ticket_cost', 0))
            invoice.transportation_fee = Decimal(request.form.get('transportation_fee', 0))
            invoice.advance_expense = Decimal(request.form.get('advance_expense', 0))
            invoice.tour_fee = Decimal(request.form.get('tour_fee', 0))
            invoice.vat = Decimal(request.form.get('vat', 0))
            invoice.withholding_tax = Decimal(request.form.get('withholding_tax', 0))
            invoice.total_tour_fee = Decimal(request.form.get('total_tour_fee', 0))
            
            db.session.commit()
            
            flash('Hong Kong Invoice updated successfully!', 'success')
            return redirect(url_for('account_report.hongkong_invoice_view', invoice_id=invoice.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating invoice: {str(e)}', 'danger')
    
    customers = Customer.query.order_by(Customer.name).all()
    return render_template('account_report/hongkong_invoice_form.html', customers=customers, invoice=invoice)


@account_report.route('/hongkong-invoice/<int:invoice_id>/delete', methods=['POST'])
@login_required
def hongkong_invoice_delete(invoice_id):
    """Delete Hong Kong invoice"""
    invoice = InvoiceHongKong.query.get_or_404(invoice_id)
    
    try:
        db.session.delete(invoice)
        db.session.commit()
        flash('Hong Kong Invoice deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting invoice: {str(e)}', 'danger')
    
    return redirect(url_for('account_report.hongkong_invoice_list'))


# ============================================
# THAI INVOICE ROUTES
# ============================================

@account_report.route('/thai-invoice')
@login_required
def thai_invoice_list():
    """List all Thai invoices"""
    invoices = InvoiceThai.query.order_by(InvoiceThai.created_at.desc()).all()
    return render_template('account_report/thai_invoice_list.html', invoices=invoices)


@account_report.route('/thai-invoice/create', methods=['GET', 'POST'])
@login_required
def thai_invoice_create():
    """Create new Thai invoice"""
    if request.method == 'POST':
        try:
            # Parse dates
            arrival_date = None
            departure_date = None
            
            if request.form.get('arrival_date'):
                arrival_date = datetime.strptime(request.form.get('arrival_date'), '%Y-%m-%d').date()
            
            if request.form.get('departure_date'):
                departure_date = datetime.strptime(request.form.get('departure_date'), '%Y-%m-%d').date()
            
            # Create new invoice
            invoice = InvoiceThai(
                customer_id=int(request.form.get('customer_id')) if request.form.get('customer_id') else None,
                cust_name=request.form.get('cust_name'),
                company_name=request.form.get('company_name'),
                company_address=request.form.get('company_address'),
                company_tel=request.form.get('company_tel'),
                company_taxid=request.form.get('company_taxid'),
                company_contact=request.form.get('company_contact'),
                total_pax=int(request.form.get('total_pax', 0)),
                arrival_date=arrival_date,
                departure_date=departure_date,
                guest_list=request.form.get('guest_list'),
                flight_info=request.form.get('flight_info'),
                air_ticket_cost=Decimal(request.form.get('air_ticket_cost', 0)),
                transportation_fee=Decimal(request.form.get('transportation_fee', 0)),
                advance_expense=Decimal(request.form.get('advance_expense', 0)),
                tour_fee=Decimal(request.form.get('tour_fee', 0)),
                vat=Decimal(request.form.get('vat', 0)),
                withholding_tax=Decimal(request.form.get('withholding_tax', 0)),
                total_tour_fee=Decimal(request.form.get('total_tour_fee', 0))
            )
            
            db.session.add(invoice)
            db.session.commit()
            
            flash('Thai Invoice created successfully!', 'success')
            return redirect(url_for('account_report.thai_invoice_view', invoice_id=invoice.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating invoice: {str(e)}', 'danger')
            return redirect(url_for('account_report.thai_invoice_create'))
    
    # GET request - show form
    customers = Customer.query.order_by(Customer.name).all()
    return render_template('account_report/thai_invoice_form.html', customers=customers, invoice=None)


@account_report.route('/thai-invoice/<int:invoice_id>')
@login_required
def thai_invoice_view(invoice_id):
    """View Thai invoice"""
    invoice = InvoiceThai.query.get_or_404(invoice_id)
    return render_template('account_report/thai_invoice_view.html', invoice=invoice)


@account_report.route('/thai-invoice/<int:invoice_id>/edit', methods=['GET', 'POST'])
@login_required
def thai_invoice_edit(invoice_id):
    """Edit Thai invoice"""
    invoice = InvoiceThai.query.get_or_404(invoice_id)
    
    if request.method == 'POST':
        try:
            # Parse dates
            if request.form.get('arrival_date'):
                invoice.arrival_date = datetime.strptime(request.form.get('arrival_date'), '%Y-%m-%d').date()
            
            if request.form.get('departure_date'):
                invoice.departure_date = datetime.strptime(request.form.get('departure_date'), '%Y-%m-%d').date()
            
            # Update fields
            invoice.customer_id = int(request.form.get('customer_id')) if request.form.get('customer_id') else None
            invoice.cust_name = request.form.get('cust_name')
            invoice.company_name = request.form.get('company_name')
            invoice.company_address = request.form.get('company_address')
            invoice.company_tel = request.form.get('company_tel')
            invoice.company_taxid = request.form.get('company_taxid')
            invoice.company_contact = request.form.get('company_contact')
            invoice.total_pax = int(request.form.get('total_pax', 0))
            invoice.guest_list = request.form.get('guest_list')
            invoice.flight_info = request.form.get('flight_info')
            invoice.air_ticket_cost = Decimal(request.form.get('air_ticket_cost', 0))
            invoice.transportation_fee = Decimal(request.form.get('transportation_fee', 0))
            invoice.advance_expense = Decimal(request.form.get('advance_expense', 0))
            invoice.tour_fee = Decimal(request.form.get('tour_fee', 0))
            invoice.vat = Decimal(request.form.get('vat', 0))
            invoice.withholding_tax = Decimal(request.form.get('withholding_tax', 0))
            invoice.total_tour_fee = Decimal(request.form.get('total_tour_fee', 0))
            
            db.session.commit()
            
            flash('Thai Invoice updated successfully!', 'success')
            return redirect(url_for('account_report.thai_invoice_view', invoice_id=invoice.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating invoice: {str(e)}', 'danger')
    
    customers = Customer.query.order_by(Customer.name).all()
    return render_template('account_report/thai_invoice_form.html', customers=customers, invoice=invoice)


@account_report.route('/thai-invoice/<int:invoice_id>/delete', methods=['POST'])
@login_required
def thai_invoice_delete(invoice_id):
    """Delete Thai invoice"""
    invoice = InvoiceThai.query.get_or_404(invoice_id)
    
    try:
        db.session.delete(invoice)
        db.session.commit()
        flash('Thai Invoice deleted successfully!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting invoice: {str(e)}', 'danger')
    
    return redirect(url_for('account_report.thai_invoice_list'))
